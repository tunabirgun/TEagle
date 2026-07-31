"""Export fidelity: the file must contain exactly what the panel showed.

TEagle's results become figures and tables in papers, so an export that drops a column, mangles a
delimiter, loses a hedge, or corrupts a unit is a scientific defect, not a cosmetic one — and one that
only a human opening a spreadsheet would notice. These tests hold the round trip: what goes in comes out.

The characters matter here. TEagle's tables carry primes (3'), en dashes in coordinate ranges (0–5146),
middots between domains (RT·INT), degree and delta signs in thermodynamics (ΔG), and percent signs. All
must survive UTF-8 encoding, delimiter escaping, and Excel's BOM convention.
"""
import builtins
import csv
import io
import os
import sys

import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
pytest.importorskip("PySide6")
_NATIVE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "app", "native")
if _NATIVE not in sys.path:
    sys.path.insert(0, _NATIVE)

from widgets import serialize_table, write_table, _csv_escape, _HAS_XLSX   # noqa: E402

# a row set that exercises every hazard at once
HEADERS = ["#", "Domain", "Coords (0-based)", "ΔG (kcal/mol)", "3′ end", "Note"]
ROWS = [
    [1, "RT·INT", "0–5146", -9.4, "TG…CA", "plain"],
    [2, "GAG", "12–99", -0.11, "AG…TT", "value, with a comma"],
    [3, "ENV", "100–200", 0.0, "—", 'quote " inside'],
    [4, "TPase", "5–10", -26.61, "n/a", "tab\tinside"],
    [5, "EN", "1–2", None, "", "newline\ninside"],
]


def _roundtrip(sep):
    text = serialize_table(HEADERS, ROWS, sep)
    return list(csv.reader(io.StringIO(text), delimiter=sep))


@pytest.mark.parametrize("sep", [",", "\t"])
def test_every_row_and_column_survives(sep):
    parsed = _roundtrip(sep)
    assert len(parsed) == len(ROWS) + 1, "row count changed between panel and file"
    assert len(parsed[0]) == len(HEADERS), "column count changed between panel and file"
    for row in parsed[1:]:
        assert len(row) == len(HEADERS), f"a row lost or gained a column: {row}"


@pytest.mark.parametrize("sep", [",", "\t"])
def test_row_order_is_preserved(sep):
    parsed = _roundtrip(sep)
    assert [r[0] for r in parsed[1:]] == ["1", "2", "3", "4", "5"]


@pytest.mark.parametrize("sep", [",", "\t"])
def test_values_containing_the_delimiter_survive(sep):
    parsed = _roundtrip(sep)
    assert parsed[2][5] == "value, with a comma"
    assert parsed[4][5] == "tab\tinside"


@pytest.mark.parametrize("sep", [",", "\t"])
def test_quotes_and_newlines_survive(sep):
    parsed = _roundtrip(sep)
    assert parsed[3][5] == 'quote " inside'
    assert parsed[5][5] == "newline\ninside"


@pytest.mark.parametrize("sep", [",", "\t"])
def test_scientific_typography_survives(sep):
    """Primes, en dashes, middots, delta and em dashes are meaning-bearing in these tables."""
    parsed = _roundtrip(sep)
    assert parsed[0][3] == "ΔG (kcal/mol)"
    assert parsed[0][4] == "3′ end"
    assert parsed[1][1] == "RT·INT"
    assert parsed[1][2] == "0–5146"
    assert parsed[3][4] == "—"


@pytest.mark.parametrize("sep", [",", "\t"])
def test_numeric_precision_is_not_rounded_away(sep):
    parsed = _roundtrip(sep)
    assert parsed[1][3] == "-9.4"
    assert parsed[2][3] == "-0.11"
    assert parsed[4][3] == "-26.61"


@pytest.mark.parametrize("sep", [",", "\t"])
def test_empty_and_none_cells_are_empty_not_the_word_none(sep):
    parsed = _roundtrip(sep)
    assert parsed[5][3] == "", "a missing value must not export as the string 'None'"
    assert parsed[5][4] == ""


def test_formula_injection_is_neutralised_without_corrupting_data():
    """CWE-1236: a cell opening with = @ or a signed number must not execute in a spreadsheet — but a
    bare strand cell '+' or '-' must stay readable."""
    assert _csv_escape("=SUM(A1:A9)", ",").startswith("'")
    assert _csv_escape("@cmd", ",").startswith("'")
    assert _csv_escape("+", ",") == "+", "a bare strand marker must not be altered"
    assert _csv_escape("-", ",") == "-"
    assert _csv_escape("-1+cmd", ",").startswith("'"), "a non-numeric signed cell is still escaped"
    # a negative number is not a formula: quoting it made every delta-G export as text
    assert _csv_escape("-9.4", ",") == "-9.4"
    assert _csv_escape("-26.61", ",") == "-26.61"
    assert _csv_escape("+1.5e-3", ",") == "+1.5e-3"


def test_csv_file_carries_the_bom_excel_needs(tmp_path):
    p = tmp_path / "t.csv"
    write_table(HEADERS, ROWS, str(p))
    raw = p.read_bytes()
    assert raw.startswith(b"\xef\xbb\xbf"), "Excel misreads UTF-8 without the BOM"
    text = raw.decode("utf-8-sig")
    assert "ΔG (kcal/mol)" in text and "3′ end" in text


def test_tsv_extension_selects_tab_separation(tmp_path):
    p = tmp_path / "t.tsv"
    write_table(HEADERS, ROWS, str(p))
    text = p.read_bytes().decode("utf-8-sig")
    assert "\t" in text.splitlines()[0]
    assert "," not in text.splitlines()[0]


@pytest.mark.skipif(not _HAS_XLSX, reason="openpyxl not available")
def test_xlsx_round_trip_matches_the_table(tmp_path):
    from openpyxl import load_workbook
    p = tmp_path / "t.xlsx"
    write_table(HEADERS, ROWS, str(p))
    ws = load_workbook(str(p)).active
    got = list(ws.values)
    assert len(got) == len(ROWS) + 1
    assert [str(h) for h in got[0]] == HEADERS
    assert got[1][1] == "RT·INT" and got[1][2] == "0–5146"
    assert got[1][3] == -9.4, "a numeric cell must stay numeric in Excel, not become text"


@pytest.mark.skipif(not _HAS_XLSX, reason="openpyxl not available")
def test_xlsx_header_is_frozen_and_bold(tmp_path):
    from openpyxl import load_workbook
    p = tmp_path / "t.xlsx"
    write_table(HEADERS, ROWS, str(p))
    ws = load_workbook(str(p)).active
    assert ws.freeze_panes == "A2"
    assert ws["A1"].font.bold


# ---------------- figure export ----------------
def _sample_model():
    """Built through figures.gv_tracks_from_rec, the same path the app uses, so the test exercises the
    real track structure rather than a hand-made one that can drift from it."""
    import figures
    rec = {"composition": {"length": 5000},
           "structural": [{"type": "LTR (terminal direct repeat)",
                           "five_prime": [100, 400], "three_prime": [4600, 4900]}],
           "domains": [{"domain": "RT", "label": "reverse transcriptase", "nt": [1000, 2000], "score": 120}],
           "orfs": [{"start": 900, "end": 3000, "strand": "+", "frame": 0, "length_aa": 700}]}
    return figures.gv_tracks_from_rec(rec)


def test_figure_svg_keeps_text_as_text_not_outlines():
    """A journal needs selectable, re-typesettable text. Outlined glyphs cannot be edited or searched."""
    import figures
    svg = figures.svg_genome(_sample_model(), {"start": 0, "end": 5000}, 900, "white")
    assert "<text" in svg, "no <text> elements — the labels were outlined into paths"
    assert "font-family" in svg


def test_figure_export_defaults_to_the_publication_palette():
    """The on-screen background is for reading; the exported file is for print. Unless the user picked a
    background explicitly, the export uses dark ink that stays legible on white paper."""
    import figures
    export_svg = figures.svg_genome(_sample_model(), {"start": 0, "end": 5000}, 900, "dark", for_export=True)
    screen_svg = figures.svg_genome(_sample_model(), {"start": 0, "end": 5000}, 900, "dark", for_export=False)
    assert export_svg != screen_svg, "export must not simply reuse the dark on-screen palette"
    assert "#E6EDF1" not in export_svg, "pale dark-theme ink would be invisible on white paper"


def test_gel_svg_is_self_contained_svg():
    import figures
    svg = figures.svg_gel({"lanes": [{"label": "P1", "amplicons": [
        {"length": 400, "on_target": True, "source": "x"}]}]}, "white")
    assert svg.strip().startswith("<svg") and svg.strip().endswith("</svg>")
    assert "<text" in svg


def test_xlsx_degrades_to_csv_when_openpyxl_is_locatable_but_broken(tmp_path, monkeypatch):
    """Deferring the openpyxl import traded an eager try/except for find_spec, which only proves the
    package is FINDABLE. A present-but-broken install would then advertise Excel export and raise on the
    first click. The writer must behave exactly as it did when openpyxl was absent: quietly write the data
    in a format that works, and stop offering the one that does not."""
    import widgets
    monkeypatch.setattr(widgets, "_XL", {}, raising=False)
    monkeypatch.setattr(widgets, "_HAS_XLSX", True, raising=False)

    real_import = builtins.__import__

    def broken(name, *a, **k):
        if name.startswith("openpyxl"):
            raise ImportError("simulated broken openpyxl")
        return real_import(name, *a, **k)

    monkeypatch.setattr(builtins, "__import__", broken)
    target = tmp_path / "t.xlsx"
    widgets.write_table(["h1", "h2"], [["a", 1]], str(target))
    monkeypatch.setattr(builtins, "__import__", real_import)

    assert not target.exists(), "no unusable .xlsx should be left behind"
    csv = tmp_path / "t.csv"
    assert csv.exists(), "the data must still be written, in a format that works"
    assert "h1" in csv.read_text(encoding="utf-8-sig")
    assert widgets._HAS_XLSX is False, "the broken format must stop being offered"
